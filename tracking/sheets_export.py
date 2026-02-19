# tracking/sheets_export.py
"""Export SQLite tracking data back to Google Sheets format.

Usage:
    python -m tracking.sheets_export                     # export to tracking/export_tracking.xlsx
    python -m tracking.sheets_export --output /tmp/out.xlsx
    python -m tracking.sheets_export --gsheets           # upload to Google Sheets
"""

from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

import json
import os
from datetime import datetime
from typing import List, Dict

from tracking.tracking_db import ApplicationTracker

# Region → sheet name
REGION_SHEETS = {
    "usa": "US",
    "non_us_international": "NonUS",
    "hk_macau_singapore": "HKMacauSingapore",
    "china_mainland": "国内",
}

STATUS_TO_FINALIZED = {
    "submitted": "Yes",
    "rec_requested": "Yes",
    "rec_submitted": "Yes",
    "long_list": "Yes",
    "short_list": "Yes",
    "offer": "Yes",
    "rejected": "Yes",
    "materials_ready": "Yes",
    "filtered_out": "not match",
    "decision_nogo": "not match",
}

STATUS_TO_NEXT_STEP = {
    "rejected": "Rejection",
    "long_list": "Long List",
    "short_list": "Short List / Interview",
    "offer": "Offer Received",
    "rec_requested": "(Optional) Ask recommendation",
    "rec_submitted": "(Optional) Ask recommendation",
}

PRIORITY_TO_FOLLOWUP = {
    "favorite": "My favorite",
    "low": "Low opportunities",
}

HEADERS = [
    "Deadline", "University", "Positions", "Application Finalized",
    "Submission Date", "Whether on time?", "Recommendation Letter Request",
    "Recommendation Letter Submitted", "Submission Recheck",
    "Next Step", "My Followup", "Notes (auto)",
]


def _fmt_date(dt_str: str) -> str:
    """Convert YYYY-MM-DD to M.D format."""
    if not dt_str:
        return ""
    try:
        d = datetime.strptime(dt_str[:10], "%Y-%m-%d")
        return f"{d.month}.{d.day}"
    except ValueError:
        return ""


def _build_row(app: dict) -> list:
    """Build a Google Sheet row from application dict."""
    status = app.get("status", "discovered")
    deadline = _fmt_date(app.get("deadline"))
    school = app.get("school", "")
    position = app.get("position", "")
    app_finalized = STATUS_TO_FINALIZED.get(status, "")
    submission_date = _fmt_date(app.get("submitted_at")) if app.get("submitted_at") else ""

    # Whether on time?
    on_time = ""
    if app.get("submitted_at") and app.get("deadline"):
        if app["submitted_at"][:10] <= app["deadline"]:
            on_time = "Yes"

    # Rec letters
    rec_letters = []
    if app.get("rec_letters"):
        try:
            rec_letters = json.loads(app["rec_letters"])
        except Exception:
            pass
    rec_request = rec_letters[0].get("source", "") if rec_letters else ""
    rec_submitted = "Yes" if any(l.get("status") == "submitted" for l in rec_letters) else ""

    next_step = STATUS_TO_NEXT_STEP.get(status, "")
    followup = PRIORITY_TO_FOLLOWUP.get(app.get("priority_tag", ""), "")

    # Auto notes
    notes_parts = []
    if app.get("fit_score"):
        notes_parts.append(f"fit={app['fit_score']:.1f}")
    if app.get("hci_strategy"):
        notes_parts.append(f"strategy={app['hci_strategy']}")
    notes = " ".join(notes_parts)

    return [
        deadline, school, position, app_finalized,
        submission_date, on_time, rec_request, rec_submitted,
        "",  # Submission Recheck
        next_step, followup, notes,
    ]


def export_to_excel(tracker: ApplicationTracker, output_path: str):
    """Export all apps to an Excel file matching the Google Sheet structure."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total = 0
    for region, sheet_name in REGION_SHEETS.items():
        apps = tracker.all_applications(region=region)
        if not apps:
            continue
        ws = wb.create_sheet(sheet_name)
        # Header row
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Data rows
        for app in apps:
            ws.append(_build_row(app))
        total += len(apps)

    wb.save(output_path)
    print(f"Exported {total} applications to {output_path}")
    return total


def export_to_gsheets(tracker: ApplicationTracker, credentials_path: str,
                      spreadsheet_id: str):
    """Upload to Google Sheets via gspread."""
    import gspread
    from google.oauth2.service_account import Credentials

    SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(spreadsheet_id)

    total = 0
    for region, sheet_name in REGION_SHEETS.items():
        apps = tracker.all_applications(region=region)
        if not apps:
            continue
        try:
            ws = spreadsheet.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=20)

        rows = [HEADERS] + [_build_row(a) for a in apps]
        ws.clear()
        ws.update(rows)
        total += len(apps)
        print(f"  {sheet_name}: {len(apps)} rows uploaded")

    print(f"Total: {total} applications uploaded")
    return total


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Export tracking DB to Google Sheets format")
    parser.add_argument("--db", default=None, help="Path to SQLite DB")
    parser.add_argument("--output", default="tracking/export_tracking.xlsx",
                        help="Output Excel path (default: tracking/export_tracking.xlsx)")
    parser.add_argument("--gsheets", action="store_true",
                        help="Upload directly to Google Sheets")
    parser.add_argument("--credentials",
                        default="google-sheets-sync/credentials/service_account.json",
                        help="Path to service account credentials")
    parser.add_argument("--spreadsheet-id",
                        default=os.getenv("SPREADSHEET_ID", ""),
                        help="Google Spreadsheet ID")
    args = parser.parse_args()

    tracker = ApplicationTracker(db_path=args.db)

    if args.gsheets:
        if not args.spreadsheet_id:
            print("Error: --spreadsheet-id required for --gsheets", file=sys.stderr)
            sys.exit(1)
        export_to_gsheets(tracker, args.credentials, args.spreadsheet_id)
    else:
        export_to_excel(tracker, args.output)


if __name__ == "__main__":
    main()
