# tracking/migration.py
"""One-shot migration: Google Sheet Excel → SQLite."""

from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

import sqlite3
import openpyxl
from tracking.tracking_db import ApplicationTracker

SHEET_REGIONS = {
    "US": "usa",
    "NonUS": "non_us_international",
    "HKMacauSingapore": "hk_macau_singapore",
    "国内": "china_mainland",
}

# 国内 has an extra "Faculty contact" column at index 10,
# pushing Next Step to 11 and Followup to 12
SHEET_COL_OFFSETS = {
    "国内": 1,
}


def _parse_deadline(raw) -> str:
    """Convert '12.1' or '1.15' to ISO date YYYY-MM-DD."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("none", ""):
        return None
    for sep in (".", "/"):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 2:
                try:
                    m, d = int(parts[0]), int(parts[1])
                    year = 2024 if m >= 8 else 2025
                    return f"{year}-{m:02d}-{d:02d}"
                except ValueError:
                    pass
    return None


def _map_row(row, col_offset: int = 0):
    """Return (status, priority_tag, rec_letters, submission_date) from sheet row."""
    app_finalized = str(row[3] or "").strip().lower()
    submission_date_raw = row[4]
    rec_request = str(row[6] or "").strip()
    rec_submitted_val = str(row[7] or "").strip().lower()
    next_step_idx = 9 + col_offset
    followup_idx = 10 + col_offset
    next_step = str(row[next_step_idx] if len(row) > next_step_idx else "").strip().lower()
    followup = str(row[followup_idx] if len(row) > followup_idx else "").strip().lower()

    submission_date = _parse_deadline(submission_date_raw) if submission_date_raw else None

    # Priority — handle "favoritate" as a typo of "favorite"
    priority_tag = None
    if "favorite" in followup or "favourite" in followup or "favoritate" in followup:
        priority_tag = "favorite"
    elif "low" in followup:
        priority_tag = "low"

    # Status
    status = "discovered"
    if app_finalized in ("not match",):
        status = "filtered_out"
    elif "expired" in app_finalized or "deadline passed" in app_finalized:
        status = "filtered_out"
    elif app_finalized in ("yes", "james_finish"):
        if submission_date:
            status = "submitted"
            # More specific post-submission overrides
            if "rejection" in next_step:
                status = "rejected"
            elif rec_submitted_val == "yes":
                status = "rec_submitted"
            elif rec_request:
                status = "rec_requested"
        else:
            status = "materials_ready"

    # Rec letters
    rec_letters = []
    if rec_request:
        letter_status = "submitted" if rec_submitted_val == "yes" else "requested"
        rec_letters.append({"source": rec_request, "status": letter_status})

    return status, priority_tag, rec_letters, submission_date


def migrate_from_excel(excel_path: str, tracker: ApplicationTracker,
                       dry_run: bool = False) -> dict:
    """Import all applications from Google Sheet Excel."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    stats = {"total_imported": 0, "skipped_empty": 0, "by_sheet": {}, "by_status": {}}

    for sheet_name, region in SHEET_REGIONS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_offset = SHEET_COL_OFFSETS.get(sheet_name, 0)
        sheet_count = 0

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            school = str(row_cells[1] or "").strip()
            if not school:
                stats["skipped_empty"] += 1
                continue

            position = str(row_cells[2] or "").strip()
            deadline = _parse_deadline(row_cells[0])

            status, priority_tag, rec_letters, submission_date = _map_row(row_cells, col_offset)

            if not dry_run:
                app_id = tracker.add_job(
                    school=school,
                    position=position,
                    region=region,
                    deadline=deadline,
                    source="google_sheet_import",
                )
                if status != "discovered":
                    tracker.update_status(app_id, status, changed_by="google_sheet_import")

                # Backfill submitted_at from submission date
                if submission_date and status in (
                    "submitted", "rec_requested", "rec_submitted", "rejected"
                ):
                    conn = sqlite3.connect(tracker.db_path)
                    conn.execute(
                        "UPDATE applications SET submitted_at=? WHERE id=?",
                        (submission_date + " 00:00:00", app_id),
                    )
                    conn.commit()
                    conn.close()

                if priority_tag:
                    tracker.set_priority(app_id, priority_tag)
                if rec_letters:
                    tracker.update_rec_letters(app_id, rec_letters)

            sheet_count += 1
            stats["total_imported"] += 1
            stats["by_status"][status] = stats["by_status"].get(status, 0) + 1

        stats["by_sheet"][sheet_name] = sheet_count

    return stats


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Migrate Google Sheet Excel → SQLite tracker")
    parser.add_argument("excel", nargs="?",
                        default="google-sheets-sync/agent_sophia_job_list.xlsx",
                        help="Path to Excel file")
    parser.add_argument("--db", default=None, help="Path to SQLite DB")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    tracker = ApplicationTracker(db_path=args.db)
    print(f"Migrating from: {args.excel}")
    print(f"Target DB: {tracker.db_path}")
    if args.dry_run:
        print("DRY RUN — no data will be written")

    stats = migrate_from_excel(args.excel, tracker, dry_run=args.dry_run)
    print(f"\nDone!")
    print(f"  Total imported: {stats['total_imported']}")
    print(f"  Skipped empty rows: {stats['skipped_empty']}")
    print(f"  By sheet: {stats['by_sheet']}")
    print(f"  By status: {stats['by_status']}")


if __name__ == "__main__":
    main()
