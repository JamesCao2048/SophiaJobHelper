#!/usr/bin/env python3
"""
全球教职每日监控系统
===========================
每日自动抓取多个学术招聘平台的 HCI / AI / CS 方向教职信息，
筛选后推荐 2-3 个最匹配的职位，并写入 Excel 追踪表。

使用方式：
    python3 faculty_monitor.py          # 运行抓取并输出今日推荐
    python3 faculty_monitor.py --all    # 输出所有抓到的职位（不仅限推荐）

依赖：requests, bs4, openpyxl (均为常见包)
如果没有 feedparser，用内置 xml.etree 解析 RSS。
"""

import os, sys, re, json, hashlib
from datetime import datetime, timedelta
from pathlib import Path
import xml.etree.ElementTree as ET

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 配置 ──────────────────────────────────────────────
KEYWORDS_HCI = [
    "human-computer interaction", "HCI", "human-AI", "human AI interaction",
    "interactive systems", "user experience", "UX research",
    "computer-supported cooperative work", "CSCW", "social computing",
    "human-centered AI", "human centered computing", "information science",
]
KEYWORDS_AI = [
    "artificial intelligence", "machine learning", "natural language processing",
    "NLP", "deep learning", "computer vision", "data science",
    "AI", "large language model", "LLM", "generative AI",
]
KEYWORDS_ALL = KEYWORDS_HCI + KEYWORDS_AI

POSITION_KEYWORDS = [
    "assistant professor", "associate professor", "professor",
    "tenure", "tenure-track", "faculty", "lecturer", "senior lecturer",
    "reader", "chair", "endowed",
    # Non-tenure track 学术教职（稳定岗位）
    "teaching professor", "research professor", "adjunct professor",
    "clinical professor", "instructor", "principal lecturer",
]

POSITION_KEYWORDS_NON_TENURE = [
    "lecturer", "senior lecturer", "teaching professor", "research professor",
    "adjunct professor", "clinical professor",
    "instructor", "principal lecturer",
    "non-tenure", "non tenure", "teaching track",
]

EXCLUDE_KEYWORDS = [
    "postdoc", "post-doc", "postdoctoral", "PhD student", "research assistant",
    "lab manager", "technician",
    # 临时性/业界导向职位
    "visiting professor", "visiting assistant professor", "research fellow",
    "professor of practice", "industry professor",
]

SCRIPT_DIR = Path(__file__).parent
EXCEL_OUTPUT = SCRIPT_DIR / "教职追踪表.xlsx"
SEEN_FILE = SCRIPT_DIR / ".seen_jobs.json"
LOG_FILE = SCRIPT_DIR / "monitor_log.txt"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/120.0.0.0 Safari/537.36"
}
TIMEOUT = 20


# ─── 工具函数 ──────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def job_id(title, url):
    raw = f"{title.strip().lower()}|{url.strip().lower()}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def load_seen():
    if SEEN_FILE.exists():
        return json.loads(SEEN_FILE.read_text(encoding="utf-8"))
    return {}


def save_seen(seen):
    SEEN_FILE.write_text(json.dumps(seen, ensure_ascii=False, indent=2), encoding="utf-8")


def matches_keywords(text, keywords):
    text_lower = text.lower()
    return any(kw.lower() in text_lower for kw in keywords)


def is_faculty_position(text):
    return matches_keywords(text, POSITION_KEYWORDS)


def is_excluded(text):
    return matches_keywords(text, EXCLUDE_KEYWORDS)


def relevance_score(title, description=""):
    combined = f"{title} {description}".lower()
    score = 0
    for kw in KEYWORDS_HCI:
        if kw.lower() in combined:
            score += 3  # HCI 方向加权
    for kw in KEYWORDS_AI:
        if kw.lower() in combined:
            score += 2
    if is_faculty_position(combined):
        score += 5
    if is_excluded(combined):
        score -= 10
    # Tenure-track 加分
    if "tenure-track" in combined or "tenure track" in combined:
        score += 3
    if "assistant professor" in combined:
        score += 2
    # Non-tenure track 同样加分（不再歧视）
    for kw in POSITION_KEYWORDS_NON_TENURE:
        if kw.lower() in combined:
            score += 2
            break  # 只加一次
    return score


def safe_get(url, **kwargs):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, **kwargs)
        resp.raise_for_status()
        return resp
    except Exception as e:
        log(f"  ⚠ 请求失败 {url}: {e}")
        return None


def parse_rss(xml_text):
    """解析 RSS/Atom feed，返回 [(title, link, description, pub_date)]"""
    items = []
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return items

    ns = {"atom": "http://www.w3.org/2005/Atom"}

    # RSS 2.0
    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        desc = (item.findtext("description") or "").strip()
        pub = (item.findtext("pubDate") or "").strip()
        if title and link:
            items.append((title, link, desc, pub))

    # Atom
    if not items:
        for entry in root.iter("{http://www.w3.org/2005/Atom}entry"):
            title = (entry.findtext("{http://www.w3.org/2005/Atom}title") or "").strip()
            link_el = entry.find("{http://www.w3.org/2005/Atom}link")
            link = link_el.get("href", "") if link_el is not None else ""
            desc = (entry.findtext("{http://www.w3.org/2005/Atom}summary") or "").strip()
            pub = (entry.findtext("{http://www.w3.org/2005/Atom}updated") or "").strip()
            if title and link:
                items.append((title, link, desc, pub))

    return items


# ─── 各平台抓取模块 ──────────────────────────────────────
def fetch_jobs_ac_uk():
    """jobs.ac.uk - 英国教职 RSS"""
    log("📡 抓取 jobs.ac.uk (RSS)...")
    jobs = []
    feeds = [
        "http://www.jobs.ac.uk/jobs/computer-sciences?format=rss",
        "http://www.jobs.ac.uk/jobs/artificial-intelligence?format=rss",
        "http://www.jobs.ac.uk/jobs/information-systems?format=rss",
    ]
    for feed_url in feeds:
        resp = safe_get(feed_url)
        if not resp:
            continue
        for title, link, desc, pub in parse_rss(resp.text):
            combined = f"{title} {desc}"
            if matches_keywords(combined, KEYWORDS_ALL) and not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "jobs.ac.uk",
                    "region": "英国/欧洲", "description": desc[:300],
                    "date_found": pub or datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title, desc),
                })
    log(f"  ✓ jobs.ac.uk 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_higheredjobs():
    """HigherEdJobs - 北美高校 RSS (可能被反爬拦截)"""
    log("📡 抓取 HigherEdJobs (RSS)...")
    jobs = []
    # Category 102 = Computer Science, 144 = Information Systems and Technology
    feeds = [
        "https://www.higheredjobs.com/search/rss.cfm?JobCat=102",
        "https://www.higheredjobs.com/search/rss.cfm?JobCat=144",
    ]
    for feed_url in feeds:
        resp = safe_get(feed_url)
        if not resp:
            continue
        # 检查返回的是否是真正的 RSS XML（而非反爬 HTML 页面）
        content = resp.text.strip()
        if content.startswith("<!DOCTYPE") or content.startswith("<html"):
            log(f"  ⚠ HigherEdJobs 返回 HTML 而非 RSS（可能被反爬拦截）")
            continue
        for title, link, desc, pub in parse_rss(content):
            combined = f"{title} {desc}"
            if (matches_keywords(combined, KEYWORDS_ALL) or is_faculty_position(title)) and not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "HigherEdJobs",
                    "region": "北美", "description": desc[:300],
                    "date_found": pub or datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title, desc),
                })
    log(f"  ✓ HigherEdJobs 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_cra():
    """CRA Career Center - JSON API"""
    log("📡 抓取 CRA Career Center (API)...")
    jobs = []
    api_headers = dict(HEADERS)
    api_headers["Accept"] = "application/json"
    search_queries = ["professor", "faculty", "HCI", "AI", "computer science"]
    seen_ids = set()
    for query in search_queries:
        try:
            resp = requests.get(
                f"https://careercenter.cra.org/api/v1/jobs?keywords={query}&resultsPerPage=25",
                headers=api_headers, timeout=TIMEOUT,
            )
            resp.raise_for_status()
            data = resp.json()
            for item in data.get("data", []):
                jid = item.get("id")
                if jid in seen_ids:
                    continue
                seen_ids.add(jid)
                title = item.get("title", "")
                url = item.get("url", "")
                desc = item.get("shortDescription", "")
                location = item.get("location", "")
                company = item.get("company", {})
                org = company.get("name", "") if isinstance(company, dict) else ""
                combined = f"{title} {desc}"
                if not is_excluded(title):
                    jobs.append({
                        "title": title, "url": url, "source": "CRA",
                        "region": f"北美 - {location}" if location else "北美",
                        "description": f"{org} - {desc[:200]}" if org else desc[:200],
                        "date_found": item.get("posted_date", datetime.now().strftime("%Y-%m-%d")),
                        "score": relevance_score(title, desc),
                    })
        except Exception as e:
            log(f"  ⚠ CRA API 查询 '{query}' 失败: {e}")
    log(f"  ✓ CRA 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_euraxess():
    """EURAXESS - 欧洲学术职位"""
    log("📡 抓取 EURAXESS...")
    jobs = []
    search_queries = [
        "HCI artificial intelligence human-computer",
        "professor computer science",
        "machine learning faculty",
    ]
    seen_urls = set()
    for kw in search_queries:
        resp = safe_get("https://euraxess.ec.europa.eu/jobs/search", params={"keywords": kw})
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("article.ecl-content-item"):
            title_el = item.select_one("h3 a, h2 a")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            link = title_el.get("href", "")
            if link and not link.startswith("http"):
                link = "https://euraxess.ec.europa.eu" + link
            if link in seen_urls:
                continue
            seen_urls.add(link)
            # 获取机构信息
            text_parts = item.get_text(" ", strip=True)
            if not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "EURAXESS",
                    "region": "欧洲", "description": text_parts[:300],
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title, text_parts),
                })
    log(f"  ✓ EURAXESS 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_chronicle():
    """Chronicle of Higher Education - 北美教职"""
    log("📡 抓取 Chronicle of Higher Education...")
    jobs = []
    search_queries = [
        "computer+science+professor",
        "HCI+faculty",
        "artificial+intelligence+professor",
    ]
    for query in search_queries:
        url = f"https://jobs.chronicle.com/searchjobs/?Keywords={query}&radialtown=&LocationId=&RadialLocation=0&CountryCode=&JobType=&SalaryFrom=&SalaryTo=&SalaryType=&PositionType="
        resp = safe_get(url)
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("div.js-clickable-area, article, div.lister__item"):
            title_el = item.select_one("h3 a, h2 a, a.js-clickable-area-link")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            link = title_el.get("href", "")
            if link and not link.startswith("http"):
                link = "https://jobs.chronicle.com" + link
            loc_el = item.select_one(".lister__meta-item, .location, .job-location")
            location = loc_el.get_text(strip=True) if loc_el else ""
            if not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "Chronicle",
                    "region": location or "北美", "description": "",
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title),
                })
    log(f"  ✓ Chronicle 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_the_unijobs():
    """Times Higher Education Unijobs"""
    log("📡 抓取 THE Unijobs...")
    jobs = []
    urls = [
        "https://www.timeshighereducation.com/unijobs/listings/computer-science/",
    ]
    for url in urls:
        resp = safe_get(url)
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("li.lister__item"):
            title_el = item.select_one("h3.lister__header a, h3 a, h2 a")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            if title.startswith("View details"):
                continue
            link = title_el.get("href", "").strip()
            if link and not link.startswith("http"):
                link = "https://www.timeshighereducation.com" + link
            combined = title
            if matches_keywords(combined, KEYWORDS_ALL) and not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "THE Unijobs",
                    "region": "全球", "description": "",
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title),
                })
    log(f"  ✓ THE Unijobs 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_inside_highered():
    """Inside Higher Ed - 北美/全球教职"""
    log("📡 抓取 Inside Higher Ed...")
    jobs = []
    search_queries = [
        "computer+science+professor",
        "HCI+faculty",
        "artificial+intelligence",
    ]
    for query in search_queries:
        url = f"https://careers.insidehighered.com/searchjobs/?Keywords={query}&radialtown=&LocationId=&RadialLocation=0"
        resp = safe_get(url)
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        for item in soup.select("div.js-clickable-area, article, div.lister__item"):
            title_el = item.select_one("h3 a, h2 a, a.js-clickable-area-link")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            link = title_el.get("href", "")
            if link and not link.startswith("http"):
                link = "https://careers.insidehighered.com" + link
            loc_el = item.select_one(".lister__meta-item, .location, .job-location")
            location = loc_el.get_text(strip=True) if loc_el else ""
            if not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "Inside Higher Ed",
                    "region": location or "北美", "description": "",
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title),
                })
    log(f"  ✓ Inside Higher Ed 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_jrecin():
    """JREC-IN Portal - 日本学术职位（英文检索）"""
    log("📡 抓取 JREC-IN (日本)...")
    jobs = []
    # JREC-IN 支持英文关键词搜索，字段 bg1=01 表示理工科
    search_queries = [
        {"keyword": "HCI human-computer interaction", "bg1": "01"},
        {"keyword": "artificial intelligence professor", "bg1": "01"},
        {"keyword": "computer science faculty", "bg1": "01"},
    ]
    seen_urls = set()
    for params in search_queries:
        resp = safe_get(
            "https://jrecin.jst.go.jp/seek/SeekJorSearch",
            params={"fn": "1", "ln": "1", "bg1": params["bg1"],
                    "sm1": "01", "keyword": params["keyword"], "lang": "1"},
        )
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        # JREC-IN 职位列表结构
        for item in soup.select("table.tbl_listtype1 tr"):
            cols = item.find_all("td")
            if len(cols) < 3:
                continue
            title_el = cols[1].find("a") if len(cols) > 1 else None
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            link = title_el.get("href", "")
            if link and not link.startswith("http"):
                link = "https://jrecin.jst.go.jp" + link
            if link in seen_urls:
                continue
            seen_urls.add(link)
            org = cols[0].get_text(strip=True) if cols else ""
            combined = f"{title} {org}"
            if matches_keywords(combined, KEYWORDS_ALL) and not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "JREC-IN",
                    "region": "日本", "description": org[:200],
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title, org),
                })
    log(f"  ✓ JREC-IN 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_academicpositions():
    """AcademicPositions.com - 全球学术职位（含亚洲/中东/欧洲）"""
    log("📡 抓取 AcademicPositions.com (全球)...")
    jobs = []
    # 尝试 RSS feed（如果存在）
    rss_urls = [
        "https://academicpositions.com/jobs/rss?field=computer-science",
        "https://academicpositions.com/jobs/rss?field=information-technology",
    ]
    for rss_url in rss_urls:
        resp = safe_get(rss_url)
        if not resp:
            continue
        content = resp.text.strip()
        if content.startswith("<") and "<item" in content:
            for title, link, desc, pub in parse_rss(content):
                combined = f"{title} {desc}"
                if not is_excluded(title):
                    jobs.append({
                        "title": title, "url": link, "source": "AcademicPositions",
                        "region": "全球", "description": desc[:300],
                        "date_found": pub or datetime.now().strftime("%Y-%m-%d"),
                        "score": relevance_score(title, desc),
                    })
    # 如果 RSS 不可用，回退到网页抓取
    if not jobs:
        search_urls = [
            "https://academicpositions.com/jobs?field=computer-science&type=professor",
            "https://academicpositions.com/jobs?field=computer-science&type=lecturer",
        ]
        seen_urls = set()
        for url in search_urls:
            resp = safe_get(url)
            if not resp:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            for item in soup.select("article.job-ad, div.job-listing, div[class*='job']"):
                title_el = item.select_one("h2 a, h3 a, a[class*='title']") or item.find("a")
                if not title_el:
                    continue
                title = title_el.get_text(strip=True)
                link = title_el.get("href", "")
                if link and not link.startswith("http"):
                    link = "https://academicpositions.com" + link
                if link in seen_urls or not link:
                    continue
                seen_urls.add(link)
                desc_el = item.select_one("p, div[class*='desc'], div[class*='summary']")
                desc = desc_el.get_text(strip=True)[:300] if desc_el else ""
                if not is_excluded(title):
                    jobs.append({
                        "title": title, "url": link, "source": "AcademicPositions",
                        "region": "全球", "description": desc,
                        "date_found": datetime.now().strftime("%Y-%m-%d"),
                        "score": relevance_score(title, desc),
                    })
    log(f"  ✓ AcademicPositions 找到 {len(jobs)} 个相关职位")
    return jobs


def fetch_seek_au():
    """Seek.com.au - 澳洲/新西兰学术职位"""
    log("📡 抓取 Seek.com.au (澳洲/新西兰)...")
    jobs = []
    search_queries = [
        "HCI professor",
        "computer science lecturer",
        "artificial intelligence faculty",
    ]
    seen_urls = set()
    for query in search_queries:
        # Seek 的搜索 URL 格式
        url = f"https://www.seek.com.au/{query.replace(' ', '-')}-jobs/in-All-Australia"
        resp = safe_get(url)
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        # Seek 使用 data-automation 属性标记职位元素
        for item in soup.select("article[data-automation='normalJob'], div[data-automation='jobCard']"):
            title_el = item.select_one("a[data-automation='jobTitle'], h3 a, h2 a")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            link = title_el.get("href", "")
            if link and not link.startswith("http"):
                link = "https://www.seek.com.au" + link
            if link in seen_urls or not link:
                continue
            seen_urls.add(link)
            org_el = item.select_one("a[data-automation='jobCompany'], span[class*='company']")
            org = org_el.get_text(strip=True) if org_el else ""
            loc_el = item.select_one("span[data-automation='jobCardLocation'], span[class*='location']")
            location = loc_el.get_text(strip=True) if loc_el else "澳洲"
            combined = f"{title} {org}"
            if (matches_keywords(combined, KEYWORDS_ALL) or is_faculty_position(title)) and not is_excluded(title):
                jobs.append({
                    "title": title, "url": link, "source": "Seek.com.au",
                    "region": f"澳洲 - {location}" if location else "澳洲",
                    "description": org[:200],
                    "date_found": datetime.now().strftime("%Y-%m-%d"),
                    "score": relevance_score(title, org),
                })
    log(f"  ✓ Seek.com.au 找到 {len(jobs)} 个相关职位")
    return jobs


# ─── 核心逻辑 ──────────────────────────────────────────
def fetch_all_jobs():
    all_jobs = []
    fetchers = [
        fetch_jobs_ac_uk,
        fetch_higheredjobs,
        fetch_cra,
        fetch_euraxess,
        fetch_chronicle,
        fetch_the_unijobs,
        fetch_inside_highered,
        # 新增：亚太 + 全球平台
        fetch_jrecin,
        fetch_academicpositions,
        fetch_seek_au,
    ]
    for fetcher in fetchers:
        try:
            jobs = fetcher()
            all_jobs.extend(jobs)
        except Exception as e:
            log(f"  ❌ {fetcher.__name__} 出错: {e}")
    return all_jobs


def deduplicate(jobs):
    seen_ids = set()
    unique = []
    for j in jobs:
        jid = job_id(j["title"], j.get("url", ""))
        if jid not in seen_ids:
            seen_ids.add(jid)
            unique.append(j)
    return unique


def filter_new_jobs(jobs, seen):
    new_jobs = []
    for j in jobs:
        jid = job_id(j["title"], j.get("url", ""))
        if jid not in seen:
            new_jobs.append(j)
            seen[jid] = {
                "title": j["title"],
                "first_seen": datetime.now().strftime("%Y-%m-%d"),
            }
    return new_jobs


def recommend_top_jobs(jobs, n=3):
    ranked = sorted(jobs, key=lambda j: j.get("score", 0), reverse=True)
    return ranked[:n]


# ─── Excel 输出 ──────────────────────────────────────────
SHEET_HEADERS = ["日期", "推荐排名", "职位名称", "学校/机构", "地区", "来源平台", "链接", "匹配度评分", "申请状态", "备注"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
STATUS_COLORS = {
    "未申请": PatternFill("solid", fgColor="FFF2CC"),
    "已申请": PatternFill("solid", fgColor="D9EAD3"),
    "面试中": PatternFill("solid", fgColor="D0E0F0"),
    "已拒": PatternFill("solid", fgColor="F4CCCC"),
}
COL_WIDTHS = [12, 10, 45, 30, 14, 18, 50, 12, 12, 25]


def init_excel():
    if EXCEL_OUTPUT.exists():
        return load_workbook(str(EXCEL_OUTPUT))
    wb = Workbook()
    ws = wb.active
    ws.title = "每日推荐"
    for col_idx, header in enumerate(SHEET_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET_HEADERS))}1"

    # 第二个 sheet：全部抓取记录
    ws2 = wb.create_sheet("全部职位")
    all_headers = ["日期", "职位名称", "来源平台", "地区", "链接", "匹配度评分", "描述摘要"]
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    col_widths_2 = [12, 45, 18, 14, 50, 12, 60]
    for col_idx, width in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.freeze_panes = "A2"

    # 第三个 sheet：平台清单（参考）
    ws3 = wb.create_sheet("监控平台")
    platform_headers = ["平台", "URL", "地区", "类型", "星级"]
    platforms = [
        ["jobs.ac.uk", "https://www.jobs.ac.uk/", "英国/欧洲", "RSS", "★★★★★"],
        ["HigherEdJobs", "https://www.higheredjobs.com/", "北美", "RSS", "★★★★"],
        ["CRA Career Center", "https://careercenter.cra.org/", "北美", "API", "★★★★★"],
        ["EURAXESS", "https://euraxess.ec.europa.eu/", "欧洲", "网页", "★★★★"],
        ["Chronicle", "https://jobs.chronicle.com/", "北美", "网页", "★★★★"],
        ["THE Unijobs", "https://www.timeshighereducation.com/unijobs/", "全球", "网页", "★★★★★"],
        ["Inside Higher Ed", "https://careers.insidehighered.com/", "北美/全球", "网页", "★★★★"],
        ["JREC-IN", "https://jrecin.jst.go.jp/", "日本", "网页", "★★★★★"],
        ["AcademicPositions", "https://academicpositions.com/", "全球(含亚洲/中东)", "RSS/网页", "★★★★"],
        ["Seek.com.au", "https://www.seek.com.au/", "澳洲/新西兰", "网页", "★★★★"],
        ["AJO (手动)", "https://academicjobsonline.org/ajo/cs", "全球", "手动", "★★★★★"],
        ["SIGCHI Jobs", "https://sigchi.org/get-involved/jobs/", "全球", "手动", "★★★★★"],
    ]
    for col_idx, h in enumerate(platform_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r, row_data in enumerate(platforms, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val).font = DATA_FONT

    wb.save(str(EXCEL_OUTPUT))
    log(f"✅ 创建新的追踪表: {EXCEL_OUTPUT}")
    return wb


def write_recommendations(wb, recommended_jobs, all_new_jobs):
    today = datetime.now().strftime("%Y-%m-%d")
    ws = wb["每日推荐"]
    start_row = ws.max_row + 1

    for rank, job in enumerate(recommended_jobs, 1):
        row_data = [
            today,
            f"Top {rank}",
            job["title"],
            "",  # 学校/机构 - 可从 description 中提取
            job.get("region", ""),
            job["source"],
            job.get("url", ""),
            job.get("score", 0),
            "未申请",
            job.get("description", "")[:80],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + rank - 1, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            if col_idx == 7 and val:  # 链接列
                cell.font = LINK_FONT
                cell.hyperlink = val
            if col_idx == 9:  # 申请状态
                cell.fill = STATUS_COLORS.get(val, PatternFill())

    # 写入全部职位 sheet
    ws2 = wb["全部职位"]
    start_row2 = ws2.max_row + 1
    for idx, job in enumerate(all_new_jobs):
        row_data = [
            today,
            job["title"],
            job["source"],
            job.get("region", ""),
            job.get("url", ""),
            job.get("score", 0),
            job.get("description", "")[:120],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=start_row2 + idx, column=col_idx, value=val)
            cell.font = DATA_FONT

    wb.save(str(EXCEL_OUTPUT))


def print_recommendations(jobs):
    today = datetime.now().strftime("%Y-%m-%d")
    print(f"\n{'='*70}")
    print(f"📋 {today} 今日教职推荐（Top {len(jobs)}）")
    print(f"{'='*70}")
    for i, job in enumerate(jobs, 1):
        print(f"\n🏆 推荐 #{i}  (匹配度: {job.get('score', 0)})")
        print(f"   职位: {job['title']}")
        print(f"   地区: {job.get('region', '未知')}")
        print(f"   来源: {job['source']}")
        print(f"   链接: {job.get('url', 'N/A')}")
        if job.get("description"):
            print(f"   摘要: {job['description'][:100]}")
    print(f"\n{'='*70}")
    print(f"💡 提示：也别忘了手动检查以下平台：")
    print(f"   • AJO: https://academicjobsonline.org/ajo/cs")
    print(f"   • SIGCHI Jobs: https://sigchi.org/get-involved/jobs/")
    print(f"   • CS Faculty Hiring Wiki: 搜索 'CS faculty hiring 2025-2026 wiki'")
    print(f"{'='*70}\n")


# ─── 主程序 ──────────────────────────────────────────
def main():
    show_all = "--all" in sys.argv
    log("🚀 开始每日教职监控...")

    # 1. 抓取所有平台
    all_jobs = fetch_all_jobs()
    all_jobs = deduplicate(all_jobs)
    log(f"📊 共抓取到 {len(all_jobs)} 个不重复职位")

    # 2. 过滤新职位
    seen = load_seen()
    new_jobs = filter_new_jobs(all_jobs, seen)
    save_seen(seen)
    log(f"🆕 其中新职位 {len(new_jobs)} 个")

    # 3. 推荐 Top 3
    if new_jobs:
        recommended = recommend_top_jobs(new_jobs, n=3)
    else:
        log("没有新职位，从所有职位中推荐...")
        recommended = recommend_top_jobs(all_jobs, n=3)

    # 4. 写入 Excel
    wb = init_excel()
    jobs_to_write = new_jobs if new_jobs else all_jobs
    write_recommendations(wb, recommended, jobs_to_write)
    log(f"📝 已写入 Excel: {EXCEL_OUTPUT}")

    # 5. 打印推荐
    print_recommendations(recommended)

    # 6. 写入追踪数据库
    try:
        import sys as _sys
        import os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        from tracking.tracking_db import ApplicationTracker
        _tracker = ApplicationTracker()
        _added = 0
        _last_id = None
        for _job in recommended:
            _last_id = _tracker.add_job(
                school=_job.get("title", "")[:80],
                position=_job.get("title", ""),
                region=_job.get("region"),
                job_url=_job.get("url"),
                source="faculty_monitor",
                monitor_score=_job.get("score"),
            )
            _added += 1
        if _added:
            log(f"📊 已将 {_added} 个推荐职位写入追踪数据库")
    except Exception as _e:
        log(f"⚠️  追踪数据库写入失败（不影响主功能）: {_e}")

    if show_all:
        print(f"\n📄 所有抓取到的职位 ({len(all_jobs)}):")
        for j in sorted(all_jobs, key=lambda x: x.get("score", 0), reverse=True):
            print(f"  [{j['score']:>3}] [{j['source']:<20}] {j['title'][:60]}")

    log("✅ 每日监控完成！")


if __name__ == "__main__":
    main()
